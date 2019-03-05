import unittest
import contextlib
from pathlib import Path, PurePath
from typing import Union, List
from io import BytesIO

import numpy as np
import openpyxl
from pyexcel_xlsx import get_data

from BioPlate import BioPlate
from BioPlate.writer.from_excel import BioPlateFromExcel, _BioPlateFromExcel

def remove_sheet(
    workbookname: str, sheetnames: Union[List, str] = ["plate_count", "plate_data"]
) -> None:
    """
        Remove given sheet name on workbook
    
    """
    wb = openpyxl.load_workbook(workbookname)
    if isinstance(sheetnames, list):
        for sheetname in sheetnames:
            wb.remove(wb[sheetname])
    else:
        wb.remove(wb[sheetnames])
    wb.active = 0
    wb.save(workbookname)
    wb.close()


class TestPlateFromExcel(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        """
        This function is run one time at the beginning of tests
        :return:
        """
        v = {
            "A[2,8]": "VC",
            "H[2,8]": "MS",
            "1-4[B,G]": ["MLR", "NT", "1.1", "1.2"],
            "E-G[8,10]": ["Val1", "Val2", "Val3"],
        }
        v1 = {
            "A[2,8]": "VC1",
            "H[2,8]": "MS1",
            "1-4[B,G]": ["MLR1", "NT1", "1.3", "1.4"],
            "E-G[8,10]": ["Val4", "Val5", "Val6"],
        }
        v2 = {
            "A[2,8]": "Top",
            "H[2,8]": "MS",
            "1-4[B,G]": ["MLR", "NT", "1.1", "1.2"],
            "E-G[8,10]": ["Val1", "Val2", "Val3"],
        }
        v3 = {
            "A[2,8]": "Bot",
            "H[2,8]": "MS1",
            "1-4[B,G]": ["MLR1", "NT1", "1.3", "1.4"],
            "E-G[8,10]": ["Val4", "Val5", "Val6"],
        }
        cls.plt = BioPlate(12, 8)
        cls.plt.set(v)
        cls.plt1 = BioPlate(12, 8)
        cls.plt1.set(v1)
        cls.stack = cls.plt + cls.plt1
        cls.Inserts = BioPlate(12, 8, inserts=True)
        cls.Inserts.top.set(v)
        cls.Inserts.bot.set(v3)
        cls.Inserts1 = BioPlate(12, 8, inserts=True)
        cls.Inserts1.bot.set(v1)
        cls.Inserts1.top.set(v2)
        cls.stacki = cls.Inserts + cls.Inserts1

    @classmethod
    def tearDownClass(cls):
        """
        This function is run one time at the end of tests
        :return:
        """
        with contextlib.suppress(FileNotFoundError):
            Path("test.xlsx").absolute().unlink()

    def setUp(self):
        """
        This function is run every time at the beginning of each test
        :return:
        """
        pass

    def tearDown(self):
        """
        This function is run every time at the end of each test
        :return:
        """
        with contextlib.suppress(FileNotFoundError):
            Path("test.xlsx").absolute().unlink()

    ###Test general on BioPlateFromExcel

    def test_BioPlateFromExcel_bp(self):
        self.plt.to_excel("test.xlsx")
        remove_sheet("test.xlsx")
        From = BioPlateFromExcel("test.xlsx")
        self.assertIsInstance(From, dict)
        np.testing.assert_array_equal(self.plt, From["plate_representation"])

    def test_BioPlateFromExcel_bpi(self):
        self.Inserts.to_excel("test.xlsx")
        remove_sheet("test.xlsx")
        From = BioPlateFromExcel("test.xlsx")
        self.assertIsInstance(From, dict)
        np.testing.assert_array_equal(self.Inserts, From["plate_representation"])

    def test_BioPlateFromExcel_bps(self):
        self.stacki.to_excel("test.xlsx")
        remove_sheet("test.xlsx")
        From = BioPlateFromExcel("test.xlsx")
        self.assertIsInstance(From, dict)
        np.testing.assert_array_equal(self.stacki, From["plate_representation"])

    def test_file_not_found(self):
        with self.assertRaises(SystemExit):
            BioPlateFromExcel("NoneFile.xlsx")

    ##Specific test on BioPlateFromExcel
    def test_is_insert_bpi(self):
        self.Inserts.to_excel("test.xlsx")
        remove_sheet("test.xlsx")
        From = _BioPlateFromExcel("test.xlsx")
        Fr = From.get_BioPlate_object()
        self.assertTrue(From.is_insert(get_data("test.xlsx")["plate_representation"]))

    def test_is_insert_bps(self):
        self.stack.to_excel("test.xlsx")
        remove_sheet("test.xlsx")
        From = _BioPlateFromExcel("test.xlsx")
        Fr = From.get_BioPlate_object()
        self.assertFalse(From.is_insert(get_data("test.xlsx")["plate_representation"]))

    def test_is_insert_bpsi(self):
        self.stacki.to_excel("test.xlsx")
        remove_sheet("test.xlsx")
        From = _BioPlateFromExcel("test.xlsx")
        Fr = From.get_BioPlate_object()
        self.assertTrue(From.is_insert(get_data("test.xlsx")["plate_representation"]))

    def test_is_stack(self):
        self.stack.to_excel("test.xlsx")
        remove_sheet("test.xlsx")
        From = _BioPlateFromExcel("test.xlsx")
        Fr = From.get_BioPlate_object()
        self.assertTrue(From.is_stack(get_data("test.xlsx")["plate_representation"]))

    # test with no header

    def test_BioPlateFromExcel_nohd_bp(self):
        self.plt.to_excel("test.xlsx", header=False)
        infos = {
            "plate_representation": {
                "row": 8,
                "column": 12,
                "stack": False,
                "type": "Plate",
            }
        }
        From = BioPlateFromExcel(
            "test.xlsx", plate_infos=infos, sheets=["plate_representation"]
        )
        self.assertIsInstance(From, dict)
        np.testing.assert_array_equal(self.plt, From["plate_representation"])

    def test_BioPlateFromExcel_nohd_bpi(self):
        self.Inserts.to_excel("test.xlsx", header=False)
        infos = {
            "plate_representation": {
                "row": 8,
                "column": 12,
                "stack": False,
                "type": "Inserts",
            }
        }
        From = BioPlateFromExcel(
            "test.xlsx", plate_infos=infos, sheets=["plate_representation"]
        )
        self.assertIsInstance(From, dict)
        np.testing.assert_array_equal(self.Inserts, From["plate_representation"])

    def test_BioPlateFromExcel_nohd_bps(self):
        self.stack.to_excel("test.xlsx", header=False)
        infos = {
            "plate_representation": {
                "row": 8,
                "column": 12,
                "stack": True,
                "type": "Plate",
            }
        }
        From = BioPlateFromExcel(
            "test.xlsx", plate_infos=infos, sheets=["plate_representation"]
        )
        self.assertIsInstance(From, dict)
        np.testing.assert_array_equal(self.stack, From["plate_representation"])

    def test_BioPlateFromExcel_nohd_bpsi(self):
        self.stacki.to_excel("test.xlsx", header=False)
        infos = {
            "plate_representation": {
                "row": 8,
                "column": 12,
                "stack": True,
                "type": "Inserts",
            }
        }
        From = BioPlateFromExcel(
            "test.xlsx", plate_infos=infos, sheets=["plate_representation"]
        )
        self.assertIsInstance(From, dict)
        np.testing.assert_array_equal(self.stacki, From["plate_representation"])


if __name__ == "__main__":
    unittest.main()
